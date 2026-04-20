"""Integration tests for async ExcelProcessor demonstrating key features."""

import asyncio
import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from src.services.excel_processor import ExcelProcessor, WorksheetProgress


@pytest.fixture
def excel_processor():
    """Create an ExcelProcessor instance for testing."""
    return ExcelProcessor()


@pytest.fixture
def complex_workbook():
    """Create a complex workbook with multiple worksheets and various features."""
    wb = Workbook()
    
    # Sheet 1: Text with formatting
    ws1 = wb.active
    ws1.title = "Formatted Text"
    ws1['A1'] = "Bold Red Text"
    ws1['A1'].font = Font(bold=True, size=14, color="FF0000")
    ws1['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws1['B1'] = "Centered Text"
    ws1['B1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws1['A2'] = "Bordered Text"
    ws1['A2'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 2: Formulas and merged cells
    ws2 = wb.create_sheet("Formulas")
    ws2['A1'] = 10
    ws2['A2'] = 20
    ws2['A3'] = "=SUM(A1:A2)"  # Formula - should be preserved
    
    # Merged cells
    ws2.merge_cells('B1:C1')
    ws2['B1'] = "Merged Cell Text"
    
    # Sheet 3: Column widths and row heights
    ws3 = wb.create_sheet("Dimensions")
    ws3.column_dimensions['A'].width = 30
    ws3.row_dimensions[1].height = 40
    ws3['A1'] = "Custom Dimensions"
    
    return wb


class TestExcelProcessorIntegration:
    """Integration tests demonstrating all key features."""
    
    @pytest.mark.asyncio
    async def test_format_preservation_during_translation(self, excel_processor, complex_workbook):
        """
        Test that all formatting is preserved when cells are updated.
        Validates: Requirements 4.5, 7.5, 12.1
        """
        worksheet = complex_workbook["Formatted Text"]
        
        # Get cells
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        
        # Store original formatting for first cell
        cell_a1 = worksheet['A1']
        original_font_bold = cell_a1.font.bold
        original_font_size = cell_a1.font.size
        original_font_color = cell_a1.font.color.rgb if cell_a1.font.color else None
        original_fill_color = cell_a1.fill.start_color.rgb if cell_a1.fill.start_color else None
        
        # Update cell with "translated" text
        await excel_processor.update_cell(cell_a1, "Translated Bold Red Text")
        
        # Verify formatting preserved
        assert cell_a1.value == "Translated Bold Red Text"
        assert cell_a1.font.bold == original_font_bold
        assert cell_a1.font.size == original_font_size
        if original_font_color:
            assert cell_a1.font.color.rgb == original_font_color
        if original_fill_color:
            assert cell_a1.fill.start_color.rgb == original_fill_color
        
        # Check alignment preservation on B1
        cell_b1 = worksheet['B1']
        original_alignment = cell_b1.alignment.horizontal
        await excel_processor.update_cell(cell_b1, "Translated Centered Text")
        assert cell_b1.alignment.horizontal == original_alignment
        
        # Check border preservation on A2
        cell_a2 = worksheet['A2']
        original_border_left = cell_a2.border.left.style if cell_a2.border.left else None
        await excel_processor.update_cell(cell_a2, "Translated Bordered Text")
        if original_border_left:
            assert cell_a2.border.left.style == original_border_left
    
    @pytest.mark.asyncio
    async def test_worksheet_property_preservation(self, excel_processor, complex_workbook):
        """
        Test that worksheet properties like column widths and row heights are preserved.
        Validates: Requirements 12.2
        """
        worksheet = complex_workbook["Dimensions"]
        
        # Store original dimensions
        original_col_width = worksheet.column_dimensions['A'].width
        original_row_height = worksheet.row_dimensions[1].height
        
        # Process the worksheet
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        for cell_data in cells:
            await excel_processor.update_cell(cell_data.cell, f"Translated: {cell_data.value}")
        
        # Verify dimensions preserved
        assert worksheet.column_dimensions['A'].width == original_col_width
        assert worksheet.row_dimensions[1].height == original_row_height
    
    @pytest.mark.asyncio
    async def test_merged_cell_preservation(self, excel_processor, complex_workbook):
        """
        Test that merged cells are preserved during processing.
        Validates: Requirements 12.4
        """
        worksheet = complex_workbook["Formulas"]
        
        # Verify merged cells exist
        assert 'B1:C1' in worksheet.merged_cells
        
        # Get cells and update
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        for cell_data in cells:
            if cell_data.value == "Merged Cell Text":
                await excel_processor.update_cell(cell_data.cell, "Translated Merged Text")
        
        # Verify merged cells still exist
        assert 'B1:C1' in worksheet.merged_cells
        assert worksheet['B1'].value == "Translated Merged Text"
    
    @pytest.mark.asyncio
    async def test_formula_preservation(self, excel_processor, complex_workbook):
        """
        Test that formulas are preserved and not translated.
        Validates: Requirements 12.5
        """
        worksheet = complex_workbook["Formulas"]
        
        # Get cells - formula cells should be skipped
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        
        # Verify formula cell is not in the list
        cell_values = [cell.value for cell in cells]
        assert "=SUM(A1:A2)" not in cell_values
        
        # Verify formula is still in the worksheet
        assert worksheet['A3'].value == "=SUM(A1:A2)"
    
    @pytest.mark.asyncio
    async def test_concurrent_worksheet_processing(self, excel_processor, complex_workbook):
        """
        Test concurrent processing of multiple worksheets.
        Validates: Requirements 4.3, 11.2
        """
        processed_count = 0
        progress_updates = []
        
        async def mock_translate_worksheet(worksheet, cells):
            """Mock translation function."""
            nonlocal processed_count
            translated = 0
            for cell_data in cells:
                # Simulate translation
                await excel_processor.update_cell(
                    cell_data.cell,
                    f"Translated: {cell_data.value}"
                )
                translated += 1
            processed_count += 1
            return translated
        
        async def track_progress(progress: WorksheetProgress):
            """Track progress updates."""
            progress_updates.append(progress)
        
        # Process all worksheets concurrently
        total_translated = await excel_processor.process_worksheets_concurrently(
            complex_workbook,
            mock_translate_worksheet,
            max_concurrency=10,
            progress_callback=track_progress
        )
        
        # Verify all worksheets were processed
        assert processed_count == len(complex_workbook.worksheets)
        assert total_translated > 0
        
        # Verify progress updates were received
        assert len(progress_updates) > 0
        
        # Verify each worksheet got progress updates
        worksheet_names = {p.worksheet_name for p in progress_updates}
        assert "Formatted Text" in worksheet_names
        assert "Formulas" in worksheet_names
        assert "Dimensions" in worksheet_names
    
    @pytest.mark.asyncio
    async def test_progress_callback_mechanism(self, excel_processor, complex_workbook):
        """
        Test that progress callbacks are called with correct information.
        Validates: Progress callback mechanism requirement
        """
        progress_history = []
        
        async def detailed_progress_callback(progress: WorksheetProgress):
            """Capture detailed progress information."""
            progress_history.append({
                'worksheet': progress.worksheet_name,
                'total': progress.cells_total,
                'processed': progress.cells_processed,
                'translated': progress.cells_translated
            })
        
        async def simple_process(worksheet, cells):
            """Simple processing function."""
            return len(cells)
        
        # Process with detailed progress tracking
        await excel_processor.process_worksheets_concurrently(
            complex_workbook,
            simple_process,
            progress_callback=detailed_progress_callback
        )
        
        # Verify progress structure
        assert len(progress_history) > 0
        for progress in progress_history:
            assert 'worksheet' in progress
            assert 'total' in progress
            assert 'processed' in progress
            assert 'translated' in progress
            assert isinstance(progress['total'], int)
            assert isinstance(progress['processed'], int)
            assert isinstance(progress['translated'], int)
    
    @pytest.mark.asyncio
    async def test_complete_workflow_with_save(self, excel_processor, complex_workbook):
        """
        Test complete workflow: load -> process -> save with format preservation.
        Validates: Complete async workflow
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            # Save original workbook
            source_path = Path(tmpdir) / "source.xlsx"
            complex_workbook.save(str(source_path))
            
            # Load workbook
            loaded_wb = await excel_processor.load_workbook(source_path)
            assert loaded_wb is not None
            
            # Process worksheets
            async def translate_cells(worksheet, cells):
                translated = 0
                for cell_data in cells:
                    await excel_processor.update_cell(
                        cell_data.cell,
                        f"VI: {cell_data.value}"
                    )
                    translated += 1
                return translated
            
            total = await excel_processor.process_worksheets_concurrently(
                loaded_wb,
                translate_cells,
                max_concurrency=10
            )
            
            assert total > 0
            
            # Save translated workbook
            output_dir = Path(tmpdir) / "output"
            output_path = await excel_processor.save_workbook(
                loaded_wb,
                source_path,
                output_dir,
                language_suffix="vi"
            )
            
            assert output_path is not None
            assert output_path.exists()
            assert output_path.name == "source_vi.xlsx"
            
            # Verify we can load the saved file
            saved_wb = await excel_processor.load_workbook(output_path)
            assert saved_wb is not None
            assert len(saved_wb.worksheets) == len(complex_workbook.worksheets)
    
    @pytest.mark.asyncio
    async def test_concurrent_limit_enforcement(self, excel_processor):
        """
        Test that concurrent processing respects the max_concurrency limit.
        Validates: Requirements 4.3, 11.2
        """
        # Create workbook with many worksheets
        wb = Workbook()
        for i in range(20):
            ws = wb.create_sheet(f"Sheet{i}")
            ws['A1'] = f"Data{i}"
        wb.remove(wb['Sheet'])  # Remove default sheet
        
        active_count = 0
        max_active = 0
        lock = asyncio.Lock()
        
        async def track_concurrency(worksheet, cells):
            """Track concurrent execution."""
            nonlocal active_count, max_active
            
            async with lock:
                active_count += 1
                max_active = max(max_active, active_count)
            
            # Simulate work
            await asyncio.sleep(0.05)
            
            async with lock:
                active_count -= 1
            
            return len(cells)
        
        # Process with limit of 5
        await excel_processor.process_worksheets_concurrently(
            wb,
            track_concurrency,
            max_concurrency=5
        )
        
        # Verify we never exceeded the limit
        assert max_active <= 5
        assert max_active > 1  # Verify we actually ran concurrently

    @pytest.mark.asyncio
    async def test_rich_text_cells_are_extracted(self, excel_processor):
        """
        Test that cells with rich text (mixed formatting) are extracted for translation.
        CellRichText objects inherit from list, not str, and must not be skipped.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "RichText"

        # Plain text cell
        ws['A1'] = "Plain text"

        # Rich text cell: mixed bold and normal text
        rich = CellRichText(
            TextBlock(InlineFont(b=True), "加粗部分"),
            "普通部分",
        )
        ws['A2'].value = rich

        # Save and reload with rich_text=True to preserve CellRichText
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "rich.xlsx"
            wb.save(str(path))

            loaded_wb = await excel_processor.load_workbook(path)
            loaded_ws = loaded_wb.active

            cells = await excel_processor.iterate_cells_in_worksheet(loaded_ws)
            values = [c.value for c in cells]

            # Both cells should be extracted (rich text converted to str)
            assert len(values) == 2
            assert "Plain text" in values
            assert "加粗部分普通部分" in values
