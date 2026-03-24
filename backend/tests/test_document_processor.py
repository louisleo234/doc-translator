"""
Tests for Document Processor Infrastructure

Tests the DocumentProcessor base class, DocumentProcessorFactory,
and ExcelDocumentProcessor adapter.
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook

from src.services.document_processor import (
    DocumentProcessor,
    DocumentType,
    TextSegment,
    ProcessingResult,
    DocumentProcessorFactory,
    apply_interleaved_mode,
    apply_output_mode,
    apply_append_mode,
)
from src.services.excel_document_processor import ExcelDocumentProcessor


class TestTextSegment:
    """Tests for TextSegment dataclass."""
    
    def test_text_segment_creation(self):
        """Test creating a TextSegment with all fields."""
        segment = TextSegment(
            id="1",
            text="Hello World",
            location="Sheet1!A1",
            metadata={"row": 1, "column": 1}
        )
        
        assert segment.id == "1"
        assert segment.text == "Hello World"
        assert segment.location == "Sheet1!A1"
        assert segment.metadata == {"row": 1, "column": 1}
    
    def test_text_segment_default_metadata(self):
        """Test TextSegment with default empty metadata."""
        segment = TextSegment(
            id="1",
            text="Test",
            location="Page 1"
        )
        
        assert segment.metadata == {}


class TestProcessingResult:
    """Tests for ProcessingResult dataclass."""
    
    def test_processing_result_success(self):
        """Test creating a successful ProcessingResult."""
        result = ProcessingResult(
            success=True,
            segments_total=100,
            segments_translated=100,
            output_path=Path("/output/file.xlsx")
        )
        
        assert result.success is True
        assert result.segments_total == 100
        assert result.segments_translated == 100
        assert result.output_path == Path("/output/file.xlsx")
        assert result.error is None
    
    def test_processing_result_failure(self):
        """Test creating a failed ProcessingResult."""
        result = ProcessingResult(
            success=False,
            segments_total=0,
            segments_translated=0,
            error="File not found"
        )
        
        assert result.success is False
        assert result.error == "File not found"
        assert result.output_path is None


class TestDocumentType:
    """Tests for DocumentType enum."""
    
    def test_document_type_values(self):
        """Test DocumentType enum values."""
        assert DocumentType.EXCEL.value == "excel"
        assert DocumentType.WORD.value == "word"
        assert DocumentType.POWERPOINT.value == "powerpoint"
        assert DocumentType.PDF.value == "pdf"
        assert DocumentType.TEXT.value == "text"
        assert DocumentType.MARKDOWN.value == "markdown"


class TestDocumentProcessorFactory:
    """Tests for DocumentProcessorFactory."""
    
    def test_factory_initialization(self):
        """Test factory initializes with empty registry."""
        factory = DocumentProcessorFactory()
        assert factory.get_supported_extensions() == []
    
    def test_register_processor(self):
        """Test registering a processor."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        
        factory.register(processor)
        
        assert '.xlsx' in factory.get_supported_extensions()
    
    def test_get_processor_for_supported_file(self):
        """Test getting processor for supported file type."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        result = factory.get_processor(Path("test.xlsx"))
        
        assert result is not None
        assert isinstance(result, ExcelDocumentProcessor)
    
    def test_get_processor_for_unsupported_file(self):
        """Test getting processor for unsupported file type."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        result = factory.get_processor(Path("test.txt"))
        
        assert result is None
    
    def test_is_supported(self):
        """Test checking if file type is supported."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        assert factory.is_supported(Path("test.xlsx")) is True
        assert factory.is_supported(Path("test.XLSX")) is True  # Case insensitive
        assert factory.is_supported(Path("test.docx")) is False
    
    def test_get_document_type(self):
        """Test getting document type for a file."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        assert factory.get_document_type(Path("test.xlsx")) == DocumentType.EXCEL
        assert factory.get_document_type(Path("test.docx")) is None


class TestExcelDocumentProcessor:
    """Tests for ExcelDocumentProcessor."""
    
    def test_supported_extensions(self):
        """Test supported extensions."""
        processor = ExcelDocumentProcessor()
        assert processor.supported_extensions == ['.xlsx']
    
    def test_document_type(self):
        """Test document type."""
        processor = ExcelDocumentProcessor()
        assert processor.document_type == DocumentType.EXCEL
    
    def test_generate_output_filename(self):
        """Test output filename generation."""
        processor = ExcelDocumentProcessor()
        
        filename = processor.generate_output_filename(Path("document.xlsx"))
        assert filename == "document_vi.xlsx"
        
        filename = processor.generate_output_filename(Path("report.xlsx"), "en")
        assert filename == "report_en.xlsx"
    
    @pytest.mark.asyncio
    async def test_extract_text_from_excel(self):
        """Test extracting text from an Excel file."""
        processor = ExcelDocumentProcessor()
        
        # Create a test Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Hello"
            ws['B1'] = "World"
            ws['A2'] = "Test"
            wb.save(f.name)
            
            file_path = Path(f.name)
        
        try:
            segments = await processor.extract_text(file_path)
            
            assert len(segments) == 3
            texts = [s.text for s in segments]
            assert "Hello" in texts
            assert "World" in texts
            assert "Test" in texts
            
            # Check metadata
            for segment in segments:
                assert "worksheet_name" in segment.metadata
                assert "row" in segment.metadata
                assert "column" in segment.metadata
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_validate_file_success(self):
        """Test validating a valid Excel file."""
        processor = ExcelDocumentProcessor()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            wb.save(f.name)
            file_path = Path(f.name)
        
        try:
            is_valid, error = await processor.validate_file(file_path)
            assert is_valid is True
            assert error is None
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_validate_file_not_found(self):
        """Test validating a non-existent file."""
        processor = ExcelDocumentProcessor()
        
        is_valid, error = await processor.validate_file(Path("/nonexistent/file.xlsx"))
        
        assert is_valid is False
        assert "not found" in error.lower()
    
    @pytest.mark.asyncio
    async def test_validate_file_wrong_extension(self):
        """Test validating a file with wrong extension."""
        processor = ExcelDocumentProcessor()
        
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as f:
            f.write(b"test")
            file_path = Path(f.name)
        
        try:
            is_valid, error = await processor.validate_file(file_path)
            assert is_valid is False
            assert "unsupported" in error.lower()
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_extract_text_skips_numeric_and_date_cells(self):
        """Test that non-text cells (numbers, dates, booleans) are skipped."""
        from datetime import datetime

        processor = ExcelDocumentProcessor()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Hello"        # string - should be extracted
            ws['B1'] = 42             # int - should be skipped
            ws['C1'] = 3.14           # float - should be skipped
            ws['A2'] = True           # bool - should be skipped
            ws['B2'] = datetime(2024, 1, 15)  # datetime - should be skipped
            ws['C2'] = "World"        # string - should be extracted
            wb.save(f.name)
            file_path = Path(f.name)

        try:
            segments = await processor.extract_text(file_path)

            texts = [s.text for s in segments]
            assert texts == ["Hello", "World"]
        finally:
            file_path.unlink()

    @pytest.mark.asyncio
    async def test_write_translated(self):
        """Test writing translated text back to Excel."""
        processor = ExcelDocumentProcessor()
        
        # Create a test Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Hello"
            ws['B1'] = "World"
            wb.save(f.name)
            input_path = Path(f.name)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)
        
        try:
            # Extract text
            segments = await processor.extract_text(input_path)
            
            # Create translations
            translations = ["Xin chào", "Thế giới"]
            
            # Write translated (with output_mode="replace" to replace original text)
            success = await processor.write_translated(
                input_path, segments, translations, output_path, output_mode="replace"
            )
            
            assert success is True
            assert output_path.exists()
            
            # Verify the translated content
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active
            assert ws['A1'].value == "Xin chào"
            assert ws['B1'].value == "Thế giới"
        finally:
            input_path.unlink()
            if output_path.exists():
                output_path.unlink()


class TestApplyInterleavedMode:
    """Tests for apply_interleaved_mode function.
    
    Validates Requirements: 4.1, 4.2, 4.3, 4.4, 4.5
    """
    
    def test_single_line_interleaving(self):
        """Test interleaving single line texts.
        
        Validates: Requirements 4.1, 4.2
        """
        original = "Hello"
        translated = "Xin chào"
        
        result = apply_interleaved_mode(original, translated)
        
        assert result == "Hello\nXin chào"
    
    def test_multi_line_equal_count_interleaving(self):
        """Test interleaving multi-line texts with equal line counts.
        
        Validates: Requirements 4.1, 4.2
        """
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nLine 3\nDòng 3"
        assert result == expected
    
    def test_original_has_more_lines(self):
        """Test interleaving when original has more lines than translated.
        
        Validates: Requirements 4.3
        """
        original = "Line 1\nLine 2\nLine 3\nLine 4"
        translated = "Dòng 1\nDòng 2"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, Dòng 2, Line 3, Line 4
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nLine 3\nLine 4"
        assert result == expected
    
    def test_translated_has_more_lines(self):
        """Test interleaving when translated has more lines than original.
        
        Validates: Requirements 4.4
        """
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2\nDòng 3\nDòng 4"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, Dòng 2, Dòng 3, Dòng 4
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nDòng 3\nDòng 4"
        assert result == expected
    
    def test_equal_texts_no_duplication(self):
        """Test that equal texts return original without duplication.
        
        Validates: Requirements 4.5
        """
        text = "Same text\nOn multiple lines"
        
        result = apply_interleaved_mode(text, text)
        
        assert result == text
    
    def test_equal_texts_with_whitespace_no_duplication(self):
        """Test that texts equal after stripping return without duplication.
        
        Validates: Requirements 4.5
        """
        original = "  Same text  "
        translated = "Same text"
        
        result = apply_interleaved_mode(original, translated)
        
        # Should return translated (no duplication)
        assert result == translated
    
    def test_empty_original(self):
        """Test interleaving with empty original text."""
        original = ""
        translated = "Dòng 1\nDòng 2"
        
        result = apply_interleaved_mode(original, translated)
        
        # Empty string splits to [''], so we get: '', Dòng 1, Dòng 2
        expected = "\nDòng 1\nDòng 2"
        assert result == expected
    
    def test_empty_translated(self):
        """Test interleaving with empty translated text."""
        original = "Line 1\nLine 2"
        translated = ""
        
        result = apply_interleaved_mode(original, translated)
        
        # Empty string splits to [''], so we get: Line 1, '', Line 2
        expected = "Line 1\n\nLine 2"
        assert result == expected
    
    def test_both_empty(self):
        """Test interleaving with both texts empty."""
        original = ""
        translated = ""
        
        result = apply_interleaved_mode(original, translated)
        
        # Both equal after strip, so return translated
        assert result == ""
    
    def test_preserves_empty_lines_in_original(self):
        """Test that empty lines in original are preserved."""
        original = "Line 1\n\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, '', Dòng 2, Line 3, Dòng 3
        expected = "Line 1\nDòng 1\n\nDòng 2\nLine 3\nDòng 3"
        assert result == expected
    
    def test_preserves_empty_lines_in_translated(self):
        """Test that empty lines in translated are preserved."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\n\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, '', Line 3, Dòng 3
        expected = "Line 1\nDòng 1\nLine 2\n\nLine 3\nDòng 3"
        assert result == expected


class TestApplyOutputMode:
    """Tests for apply_output_mode dispatcher function.
    
    Validates: Requirements 4.6
    """
    
    def test_replace_mode_returns_translated_only(self):
        """Test replace mode returns translated text only.

        Validates: Requirements 4.6
        """
        original = "Hello World"
        translated = "Xin chào Thế giới"

        result = apply_output_mode(original, translated, "replace")

        assert result == translated

    def test_append_mode_appends_translated_after_original(self):
        """Test append mode appends translated text after original.

        Validates: Requirements 4.6
        """
        original = "Hello World"
        translated = "Xin chào Thế giới"

        result = apply_output_mode(original, translated, "append")

        expected = f"{original}\n{translated}"
        assert result == expected

    def test_interleaved_mode_interleaves_lines(self):
        """Test interleaved mode interleaves original and translated lines.

        Validates: Requirements 4.6
        """
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"

        result = apply_output_mode(original, translated, "interleaved")

        expected = "Line 1\nDòng 1\nLine 2\nDòng 2"
        assert result == expected

    def test_replace_mode_with_multiline_text(self):
        """Test replace mode with multi-line text."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"

        result = apply_output_mode(original, translated, "replace")

        assert result == translated

    def test_append_mode_with_multiline_text(self):
        """Test append mode with multi-line text."""
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"

        result = apply_output_mode(original, translated, "append")

        expected = f"{original}\n{translated}"
        assert result == expected

    def test_replace_mode_with_empty_translated(self):
        """Test replace mode with empty translated text."""
        original = "Hello World"
        translated = ""

        result = apply_output_mode(original, translated, "replace")

        assert result == ""

    def test_append_mode_with_equal_texts(self):
        """Test append mode when original and translated are equal (no duplication)."""
        text = "Same text"

        result = apply_output_mode(text, text, "append")

        # apply_append_mode should avoid duplication when texts are equal
        assert result == text

    def test_interleaved_mode_with_equal_texts(self):
        """Test interleaved mode when original and translated are equal (no duplication)."""
        text = "Same text\nOn multiple lines"

        result = apply_output_mode(text, text, "interleaved")

        # apply_interleaved_mode should avoid duplication when texts are equal
        assert result == text

    def test_interleaved_mode_with_unequal_line_counts(self):
        """Test interleaved mode handles unequal line counts correctly."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1"

        result = apply_output_mode(original, translated, "interleaved")

        # Expected: Line 1, Dòng 1, Line 2, Line 3
        expected = "Line 1\nDòng 1\nLine 2\nLine 3"
        assert result == expected

    def test_replace_mode_preserves_special_characters(self):
        """Test replace mode preserves special characters in translated text."""
        original = "Hello"
        translated = "Xin chào! 你好 🌍"

        result = apply_output_mode(original, translated, "replace")

        assert result == translated

    def test_append_mode_preserves_whitespace(self):
        """Test append mode preserves whitespace in both texts."""
        original = "  Hello  "
        translated = "  Xin chào  "

        result = apply_output_mode(original, translated, "append")

        expected = f"{original}\n{translated}"
        assert result == expected

    def test_unknown_mode_falls_back_to_replace(self):
        """Test that an unknown output_mode falls back to replace behavior."""
        original = "Hello World"
        translated = "Xin chào Thế giới"

        result = apply_output_mode(original, translated, "unknown_mode")

        assert result == translated

    def test_default_mode_is_replace(self):
        """Test that the default output_mode is replace."""
        original = "Hello World"
        translated = "Xin chào Thế giới"

        result = apply_output_mode(original, translated)

        assert result == translated


class TestApplyAppendMode:
    """Tests for the apply_append_mode function."""

    def test_appends_translated_after_original(self):
        """Test that translated text is appended after original."""
        original = "Hello"
        translated = "Xin chào"

        result = apply_append_mode(original, translated)

        assert result == "Hello\nXin chào"

    def test_no_duplication_when_texts_equal(self):
        """Test that equal texts are not duplicated."""
        text = "Same text"

        result = apply_append_mode(text, text)

        assert result == text

    def test_no_duplication_when_texts_equal_after_strip(self):
        """Test that texts equal after stripping are not duplicated."""
        original = "  Hello  "
        translated = "Hello"

        result = apply_append_mode(original, translated)

        assert result == translated

    def test_multiline_append(self):
        """Test append with multiline text."""
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"

        result = apply_append_mode(original, translated)

        assert result == "Line 1\nLine 2\nDòng 1\nDòng 2"


class TestOutputModeInTranslationJob:
    """Tests for output_mode field in TranslationJob model."""

    def test_default_output_mode_is_replace(self):
        """Test that TranslationJob defaults output_mode to 'replace'."""
        from src.models.job import TranslationJob, JobStatus

        job = TranslationJob(
            status=JobStatus.PENDING,
            files_total=1,
            file_ids=["file-1"]
        )

        assert job.output_mode == "replace"

    def test_output_mode_set_to_append(self):
        """Test that output_mode can be set to 'append'."""
        from src.models.job import TranslationJob, JobStatus

        job = TranslationJob(
            status=JobStatus.PENDING,
            files_total=1,
            file_ids=["file-1"],
            output_mode="append"
        )

        assert job.output_mode == "append"

    def test_output_mode_set_to_interleaved(self):
        """Test that output_mode can be set to 'interleaved'."""
        from src.models.job import TranslationJob, JobStatus

        job = TranslationJob(
            status=JobStatus.PENDING,
            files_total=1,
            file_ids=["file-1"],
            output_mode="interleaved"
        )

        assert job.output_mode == "interleaved"


class TestOutputModeInJobManager:
    """Tests for output_mode propagation through JobManager."""

    @pytest.fixture
    def job_manager(self):
        from unittest.mock import AsyncMock, Mock
        from src.services.job_manager import JobManager
        mock_store = Mock()
        mock_store.save_job = AsyncMock()
        return JobManager(job_store=mock_store)

    @pytest.fixture
    def language_pair(self):
        from src.models.job import LanguagePair
        return LanguagePair(
            id="zh-vi",
            source_language="Chinese",
            target_language="Vietnamese",
            source_language_code="zh",
            target_language_code="vi"
        )

    async def test_create_job_default_output_mode(self, job_manager, language_pair):
        """Test that create_job defaults output_mode to 'replace'."""
        job = await job_manager.create_job(["file-1"], language_pair)

        assert job.output_mode == "replace"

    async def test_create_job_with_append_mode(self, job_manager, language_pair):
        """Test that create_job propagates 'append' output_mode."""
        job = await job_manager.create_job(
            ["file-1"], language_pair, output_mode="append"
        )

        assert job.output_mode == "append"

    async def test_create_job_with_interleaved_mode(self, job_manager, language_pair):
        """Test that create_job propagates 'interleaved' output_mode."""
        job = await job_manager.create_job(
            ["file-1"], language_pair, output_mode="interleaved"
        )

        assert job.output_mode == "interleaved"


class TestOutputModeValidationInResolver:
    """Tests for output_mode validation in the GraphQL resolver."""

    async def test_invalid_output_mode_raises_validation_error(self):
        """Test that invalid output_mode raises ValidationError."""
        from unittest.mock import Mock
        from src.graphql.resolvers import resolve_create_translation_job

        info = Mock()
        info.context = {
            "request": Mock(headers={"Authorization": "Bearer test-token"}),
            "resolver_context": Mock(),
            "current_user": Mock(username="testuser"),
        }

        with pytest.raises(Exception) as exc_info:
            await resolve_create_translation_job(
                info=info,
                file_ids=["file-1"],
                language_pair_id="lp-1",
                output_mode="invalid_mode"
            )

        assert "Invalid output_mode" in str(exc_info.value)

    async def test_empty_string_output_mode_raises_validation_error(self):
        """Test that empty string output_mode raises ValidationError."""
        from unittest.mock import Mock
        from src.graphql.resolvers import resolve_create_translation_job

        info = Mock()
        info.context = {
            "request": Mock(headers={"Authorization": "Bearer test-token"}),
            "resolver_context": Mock(),
            "current_user": Mock(username="testuser"),
        }

        with pytest.raises(Exception) as exc_info:
            await resolve_create_translation_job(
                info=info,
                file_ids=["file-1"],
                language_pair_id="lp-1",
                output_mode=""
            )

        assert "Invalid output_mode" in str(exc_info.value)

    async def test_case_sensitive_output_mode_validation(self):
        """Test that output_mode validation is case-sensitive."""
        from unittest.mock import Mock
        from src.graphql.resolvers import resolve_create_translation_job

        info = Mock()
        info.context = {
            "request": Mock(headers={"Authorization": "Bearer test-token"}),
            "resolver_context": Mock(),
            "current_user": Mock(username="testuser"),
        }

        for invalid in ["Replace", "APPEND", "Interleaved"]:
            with pytest.raises(Exception) as exc_info:
                await resolve_create_translation_job(
                    info=info,
                    file_ids=["file-1"],
                    language_pair_id="lp-1",
                    output_mode=invalid
                )

            assert "Invalid output_mode" in str(exc_info.value)
