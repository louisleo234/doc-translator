"""
Excel Processor Module (Async Version)

Handles Excel file reading, cell iteration, translation updates, and format preservation
with async support for concurrent worksheet processing.
"""

import asyncio
import logging
from copy import copy
from pathlib import Path
from typing import Optional, List, Callable, Any, Awaitable
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


@dataclass
class CellData:
    """Data structure for cell information"""
    worksheet_name: str
    row: int
    column: int
    value: Any
    has_formula: bool
    cell: Cell  # Keep reference to actual cell for updates


@dataclass
class WorksheetProgress:
    """Progress information for a worksheet"""
    worksheet_name: str
    cells_total: int
    cells_processed: int
    cells_translated: int


class ExcelProcessor:
    """Async processor for Excel file operations with format preservation"""
    
    def __init__(self, logger: Optional[logging.Logger] = None):
        """
        Initialize the Excel processor.
        
        Args:
            logger: Logger instance for logging operations
        """
        self.logger = logger or logging.getLogger(__name__)
    
    async def load_workbook(self, file_path: Path) -> Optional[Workbook]:
        """
        Load an Excel workbook with format preservation (async wrapper).
        
        Uses openpyxl with data_only=False to preserve formulas and formatting.
        Note: keep_vba is disabled to prevent corruption issues with MS Excel.
        
        Args:
            file_path: Path to the Excel file to load
            
        Returns:
            Workbook object if successful, None if loading fails
        """
        try:
            self.logger.info(f"Loading workbook: {file_path.name}")
            
            # Run blocking I/O operation in thread pool
            workbook = await asyncio.to_thread(
                load_workbook,
                filename=str(file_path),
                keep_links=True,     # Preserve external links
                data_only=False,     # Keep formulas (not just calculated values)
                rich_text=True       # Preserve rich text formatting
            )
            
            self.logger.info(f"Workbook loaded successfully: {len(workbook.worksheets)} worksheet(s)")
            return workbook
            
        except FileNotFoundError as e:
            self.logger.error(f"File not found: {file_path} - {str(e)}")
            return None
        except PermissionError as e:
            self.logger.error(f"Permission denied accessing file: {file_path} - {str(e)}")
            return None
        except Exception as e:
            self.logger.error(f"Failed to load workbook {file_path.name}: {type(e).__name__} - {str(e)}")
            return None

    async def iterate_cells_in_worksheet(
        self,
        worksheet: Worksheet
    ) -> List[CellData]:
        """
        Iterate through all cells in a single worksheet.
        
        Returns list of CellData for each non-empty cell. Skips cells with only whitespace.
        Preserves formula cells without translating the formula text itself.
        
        Args:
            worksheet: The worksheet to iterate through
            
        Returns:
            List of CellData objects for each cell containing text
        """
        worksheet_name = worksheet.title
        self.logger.debug(f"Processing worksheet: {worksheet_name}")
        
        cells = []
        
        # Run iteration in thread pool to avoid blocking
        def _iterate():
            result = []
            for row in worksheet.iter_rows():
                for cell in row:
                    # Skip empty cells
                    if cell.value is None:
                        continue
                    
                    # Check if cell has a formula
                    has_formula = isinstance(cell.value, str) and cell.value.startswith('=')
                    
                    # For formula cells, we don't translate the formula itself
                    # We preserve the formula as-is
                    if has_formula:
                        self.logger.debug(f"Skipping formula cell at {worksheet_name}!{cell.coordinate}")
                        continue

                    # Skip non-text cells (numbers, dates, booleans)
                    if not isinstance(cell.value, str):
                        continue

                    cell_text = cell.value
                    
                    # Skip cells with only whitespace
                    if not cell_text.strip():
                        continue
                    
                    # Add cell data for translation
                    result.append(CellData(
                        worksheet_name=worksheet_name,
                        row=cell.row,
                        column=cell.column,
                        value=cell_text,
                        has_formula=has_formula,
                        cell=cell  # Keep reference for later updates
                    ))
            return result
        
        cells = await asyncio.to_thread(_iterate)
        return cells

    async def process_worksheets_concurrently(
        self,
        workbook: Workbook,
        process_func: Callable[[Worksheet, List[CellData]], Awaitable[int]],
        max_concurrency: int = 10,
        progress_callback: Optional[Callable[[WorksheetProgress], Awaitable[None]]] = None
    ) -> int:
        """
        Process multiple worksheets concurrently with semaphore-based limiting.
        
        Args:
            workbook: The workbook containing worksheets to process
            process_func: Async function to process each worksheet (receives worksheet and cells)
            max_concurrency: Maximum number of worksheets to process concurrently
            progress_callback: Optional callback for progress updates
            
        Returns:
            Total number of cells translated across all worksheets
        """
        semaphore = asyncio.Semaphore(max_concurrency)
        total_translated = 0
        
        async def _process_worksheet(worksheet: Worksheet) -> int:
            async with semaphore:
                # Get all cells in this worksheet
                cells = await self.iterate_cells_in_worksheet(worksheet)
                
                # Report initial progress
                if progress_callback:
                    await progress_callback(WorksheetProgress(
                        worksheet_name=worksheet.title,
                        cells_total=len(cells),
                        cells_processed=0,
                        cells_translated=0
                    ))
                
                # Process the worksheet
                translated_count = await process_func(worksheet, cells)
                
                # Report completion
                if progress_callback:
                    await progress_callback(WorksheetProgress(
                        worksheet_name=worksheet.title,
                        cells_total=len(cells),
                        cells_processed=len(cells),
                        cells_translated=translated_count
                    ))
                
                return translated_count
        
        # Create tasks for all worksheets
        tasks = [_process_worksheet(ws) for ws in workbook.worksheets]
        
        # Execute concurrently and gather results
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Sum up translated cells (handle exceptions)
        for result in results:
            if isinstance(result, Exception):
                self.logger.error(f"Worksheet processing error: {result}")
            else:
                total_translated += result
        
        return total_translated

    async def update_cell(self, cell: Cell, translated_text: str) -> None:
        """
        Update a cell with translated text while preserving all formatting (async wrapper).
        
        Preserves:
        - Font properties (family, size, color, bold, italic, underline)
        - Fill properties (background color, pattern)
        - Border properties (all sides)
        - Alignment properties (horizontal, vertical, wrap text)
        - Number format
        - Data validation rules
        
        Args:
            cell: The cell to update
            translated_text: The translated text to set
        """
        def _update():
            # Store original formatting before updating value
            original_font = copy(cell.font) if cell.font else None
            original_fill = copy(cell.fill) if cell.fill else None
            original_border = copy(cell.border) if cell.border else None
            original_alignment = copy(cell.alignment) if cell.alignment else None
            original_number_format = cell.number_format
            
            # Update cell value with translation
            cell.value = translated_text
            
            # Restore formatting
            if original_font:
                cell.font = original_font
            if original_fill:
                cell.fill = original_fill
            if original_border:
                cell.border = original_border
            if original_alignment:
                cell.alignment = original_alignment
            if original_number_format:
                cell.number_format = original_number_format
            
            self.logger.debug(f"Updated cell {cell.coordinate} with translation (format preserved)")
        
        await asyncio.to_thread(_update)

    async def save_workbook(
        self,
        workbook: Workbook,
        source_file_path: Path,
        output_dir: Path,
        language_suffix: str = "vi"
    ) -> Optional[Path]:
        """
        Save the translated workbook to the output directory with format preservation.
        
        Generates output filename with language suffix before extension.
        Preserves column widths, row heights, merged cells, and worksheet properties.
        Overwrites existing files if present.
        
        Args:
            workbook: The workbook to save
            source_file_path: Original file path (used to generate output filename)
            output_dir: Directory to save the translated file
            language_suffix: Language code suffix (default: "vi")
            
        Returns:
            Path to the saved file if successful, None otherwise
        """
        try:
            # Generate output filename with language suffix
            source_name = source_file_path.stem  # Filename without extension
            source_ext = source_file_path.suffix  # Extension (e.g., .xlsx)
            output_filename = f"{source_name}_{language_suffix}{source_ext}"
            output_path = output_dir / output_filename
            
            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)
            
            self.logger.info(f"Saving translated workbook to: {output_path.name}")
            
            # Save the workbook in thread pool (blocking I/O)
            await asyncio.to_thread(workbook.save, str(output_path))
            
            self.logger.info(f"Workbook saved successfully: {output_path.name}")
            return output_path
            
        except PermissionError as e:
            self.logger.error(f"Permission denied saving file to {output_path}: {str(e)}")
            return None
        except OSError as e:
            self.logger.error(f"OS error saving file to {output_path}: {str(e)}")
            return None
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {type(e).__name__} - {str(e)}")
            return None
